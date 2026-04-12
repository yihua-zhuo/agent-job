"""
文件处理辅助工具
"""
import csv
import openpyxl
from io import BytesIO, StringIO
from typing import List, Dict


class FileHelper:
    """文件处理辅助工具"""

    @staticmethod
    def read_csv(content: bytes) -> List[Dict]:
        """读取CSV文件
        
        Args:
            content: CSV文件字节内容
            
        Returns:
            List[Dict]: CSV数据列表，每项为字典
        """
        result = []
        try:
            # 将bytes转换为字符串
            text = content.decode('utf-8')
            if content.startswith(b'\xef\xbb\xbf'):  # BOM处理
                text = content.decode('utf-8-sig')
            
            reader = csv.DictReader(StringIO(text))
            for row in reader:
                # 清理空值
                cleaned_row = {k: v for k, v in row.items() if v is not None}
                result.append(cleaned_row)
        except UnicodeDecodeError:
            # 尝试其他编码
            try:
                text = content.decode('gbk')
                reader = csv.DictReader(StringIO(text))
                for row in reader:
                    cleaned_row = {k: v for k, v in row.items() if v is not None}
                    result.append(cleaned_row)
            except UnicodeDecodeError:
                text = content.decode('latin-1', errors='ignore')
                reader = csv.DictReader(StringIO(text))
                for row in reader:
                    cleaned_row = {k: v for k, v in row.items() if v is not None}
                    result.append(cleaned_row)
        
        return result

    @staticmethod
    def write_csv(data: List[Dict], columns: List[str]) -> bytes:
        """写入CSV文件
        
        Args:
            data: 数据列表
            columns: 列名列表
            
        Returns:
            bytes: CSV文件内容
        """
        if not data:
            return b""
        
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue().encode('utf-8-sig')  # 添加BOM以便Excel正确识别中文

    @staticmethod
    def read_excel(content: bytes) -> List[Dict]:
        """读取Excel文件
        
        Args:
            content: Excel文件字节内容
            
        Returns:
            List[Dict]: Excel数据列表
        """
        result = []
        try:
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 读取数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue
                    
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = value
                
                if row_dict:
                    result.append(row_dict)
            
            wb.close()
        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")
        
        return result

    @staticmethod
    def write_excel(data: List[Dict], sheet_name: str = "Sheet1") -> bytes:
        """写入Excel文件
        
        Args:
            data: 数据列表
            sheet_name: 工作表名称
            
        Returns:
            bytes: Excel文件内容
        """
        if not data:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 获取所有列名
        columns = list(data[0].keys())
        
        # 写入表头
        for col_idx, column in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=column)
        
        # 写入数据
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, column in enumerate(columns, start=1):
                value = row_data.get(column, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col_idx, column in enumerate(columns, start=1):
            max_length = len(str(column))
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def detect_file_format(filename: str) -> str:
        """根据文件名检测格式
        
        Args:
            filename: 文件名
            
        Returns:
            str: 文件格式 (csv, excel, json, pdf)
        """
        if not filename:
            return ""
        
        filename_lower = filename.lower()
        
        if filename_lower.endswith('.csv'):
            return "csv"
        elif filename_lower.endswith(('.xlsx', '.xls')):
            return "excel"
        elif filename_lower.endswith('.json'):
            return "json"
        elif filename_lower.endswith('.pdf'):
            return "pdf"
        else:
            return ""
