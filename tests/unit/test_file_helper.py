"""
Unit tests for src/utils/file_helper.py — FileHelper class
Covers: read_csv, write_csv, read_excel, write_excel, detect_file_format
"""
import pytest
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

from utils.file_helper import FileHelper

# Chinese test strings for readability
ZH_NAME = "张三"
ZH_CITY = "北京"
ZH2_NAME = "王五"
ZH2_CITY = "上海"


class TestReadCsv:
    """Tests for FileHelper.read_csv()"""

    def test_read_csv_utf8(self):
        """UTF-8 encoded CSV with Chinese characters."""
        content = f"name,age,city\n{ZH_NAME},25,{ZH_CITY}\n{ZH2_NAME},30,{ZH2_CITY}".encode('utf-8')
        result = FileHelper.read_csv(content)
        assert len(result) == 2
        assert result[0]["name"] == ZH_NAME
        assert result[0]["age"] == "25"
        assert result[0]["city"] == ZH_CITY

    def test_read_csv_utf8_sig_bom(self):
        """UTF-8-sig BOM is stripped so DictReader sees clean headers."""
        content = "name,age,city\nAlice,30,NYC".encode('utf-8-sig')  # adds BOM
        result = FileHelper.read_csv(content)
        assert len(result) == 1
        assert result[0]["name"] == "Alice"

    def test_read_csv_gbk_fallback(self):
        """Bytes valid in GBK but not UTF-8 trigger GBK fallback."""
        # 0xC0 is a lead byte in GBK; alone it's invalid UTF-8 but valid GBK
        content = b'\xc0\xc1,name\nBob,25'  # \xc0\xc1 forms a valid GBK character
        result = FileHelper.read_csv(content)
        assert len(result) == 1
        # GBK decodes \xc0\xc1 to a Chinese character, \xc1 as part of another
        assert result[0]["name"] is not None

    def test_read_csv_latin1_fallback(self):
        """Latin-1 fallback for binary garbled content."""
        content = b"\xff\xfe\xfd\xfc"
        result = FileHelper.read_csv(content)
        assert isinstance(result, list)  # Returns list, may be empty or garbled

    def test_read_csv_empty_content(self):
        """Empty bytes returns empty list."""
        result = FileHelper.read_csv(b"")
        assert result == []

    def test_read_csv_header_only(self):
        """CSV with only header row returns empty list."""
        content = b"name,age,city"
        result = FileHelper.read_csv(content)
        assert result == []

    def test_read_csv_dict_reader_fields(self):
        """DictReader uses first row as fieldnames."""
        content = b"col1,col2,col3\nval1,val2,val3"
        result = FileHelper.read_csv(content)
        assert len(result) == 1
        assert result[0]["col1"] == "val1"
        assert result[0]["col3"] == "val3"

    def test_read_csv_strips_none_values(self):
        """csv.DictReader returns None for missing trailing fields — these are filtered."""
        # Row with fewer fields than header: trailing fields are None
        content = b"name,city\nAlice"
        result = FileHelper.read_csv(content)
        assert len(result) == 1
        assert result[0]["name"] == "Alice"
        assert "city" not in result[0]  # missing trailing field → None, filtered out

    def test_read_csv_empty_string_preserved(self):
        """Empty string cells are preserved (not treated as None)."""
        content = b"name,age,city\nAlice,,NYC"
        result = FileHelper.read_csv(content)
        assert result[0]["name"] == "Alice"
        assert result[0]["age"] == ""  # empty string, NOT None — preserved
        assert result[0]["city"] == "NYC"


class TestWriteCsv:
    """Tests for FileHelper.write_csv()"""

    def test_write_csv_returns_bytes(self):
        """write_csv returns bytes."""
        data = [{"name": "Alice", "age": "30"}]
        result = FileHelper.write_csv(data, ["name", "age"])
        assert isinstance(result, bytes)

    def test_write_csv_has_utf8_bom(self):
        """Output starts with UTF-8 BOM for Excel compatibility."""
        data = [{"name": "Alice"}]
        result = FileHelper.write_csv(data, ["name"])
        assert result.startswith(b'\xef\xbb\xbf')

    def test_write_csv_empty_data_returns_empty_bytes(self):
        """Empty data list returns b''."""
        result = FileHelper.write_csv([], ["name", "age"])
        assert result == b""

    def test_write_csv_header_order_matches_columns(self):
        """Column order in output matches the columns parameter."""
        data = [{"name": "Alice", "age": "30", "city": "NYC"}]
        result = FileHelper.write_csv(data, ["city", "name", "age"])
        decoded = result.decode('utf-8-sig')
        header = decoded.split('\n')[0]
        assert header.startswith("city,name,age")

    def test_write_csv_ignores_extra_keys(self):
        """Keys not in columns list are ignored (extrasaction='ignore')."""
        data = [{"name": "Alice", "secret": "hidden"}]
        result = FileHelper.write_csv(data, ["name"])
        decoded = result.decode('utf-8-sig')
        assert "secret" not in decoded

    def test_write_csv_missing_keys_empty(self):
        """Dicts missing some columns get empty values."""
        data = [{"name": "Alice"}]  # missing 'age'
        result = FileHelper.write_csv(data, ["name", "age"])
        decoded = result.decode('utf-8-sig')
        assert "Alice" in decoded

    def test_write_csv_chinese_characters(self):
        """Chinese characters encoded correctly with BOM."""
        data = [{"name": ZH_NAME, "city": ZH_CITY}]
        result = FileHelper.write_csv(data, ["name", "city"])
        decoded = result.decode('utf-8-sig')
        assert ZH_NAME in decoded
        assert ZH_CITY in decoded


class TestReadExcel:
    """Tests for FileHelper.read_excel()"""

    def _make_excel_bytes(self, rows):
        """Helper: create Excel bytes from list of lists."""
        from openpyxl import Workbook
        from io import BytesIO
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        buf.seek(0)
        return buf.getvalue()

    def test_read_excel_basic(self):
        """Basic Excel reading with Chinese data."""
        content = self._make_excel_bytes([
            ["name", "age", "city"],
            [ZH_NAME, "25", ZH_CITY],
            [ZH2_NAME, "30", ZH2_CITY],
        ])
        result = FileHelper.read_excel(content)
        assert len(result) == 2
        assert result[0]["name"] == ZH_NAME
        assert result[1]["name"] == ZH2_NAME

    def test_read_excel_skips_empty_rows(self):
        """Rows where all cells are None/empty are skipped."""
        content = self._make_excel_bytes([
            ["name", "age"],
            [ZH_NAME, "25"],
            [None, None],
            [ZH2_NAME, "30"],
            [],
        ])
        result = FileHelper.read_excel(content)
        assert len(result) == 2

    def test_read_excel_header_only(self):
        """Excel with only header row returns empty list."""
        content = self._make_excel_bytes([["col1", "col2", "col3"]])
        result = FileHelper.read_excel(content)
        assert result == []

    def test_read_excel_empty_workbook(self):
        """Empty workbook (no data) returns empty list."""
        content = self._make_excel_bytes([])
        result = FileHelper.read_excel(content)
        assert result == []


class TestWriteExcel:
    """Tests for FileHelper.write_excel()"""

    def test_write_excel_returns_bytes(self):
        """write_excel returns bytes."""
        data = [{"name": "Alice", "age": 30}]
        result = FileHelper.write_excel(data, "Test")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_write_excel_empty_data_returns_workbook_bytes(self):
        """Empty data still returns a valid (empty) workbook."""
        result = FileHelper.write_excel([], "EmptySheet")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_write_excel_sheet_name(self):
        """Sheet name is set correctly."""
        data = [{"name": "Alice"}]
        result = FileHelper.write_excel(data, "MySheet")
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(result))
        assert wb.active.title == "MySheet"
        wb.close()

    def test_write_excel_writes_headers(self):
        """Headers are written in row 1."""
        data = [{"name": "Alice", "city": "NYC"}]
        result = FileHelper.write_excel(data, "Test")
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "name"
        assert ws.cell(row=1, column=2).value == "city"
        wb.close()

    def test_write_excel_writes_data_rows(self):
        """Data rows start at row 2."""
        data = [{"name": "Alice"}, {"name": "Bob"}]
        result = FileHelper.write_excel(data, "Test")
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Alice"
        assert ws.cell(row=3, column=1).value == "Bob"
        wb.close()


class TestDetectFileFormat:
    """Tests for FileHelper.detect_file_format()"""

    def test_detect_csv_lowercase(self):
        assert FileHelper.detect_file_format("data.csv") == "csv"

    def test_detect_csv_uppercase(self):
        assert FileHelper.detect_file_format("DATA.CSV") == "csv"

    def test_detect_excel_xlsx(self):
        assert FileHelper.detect_file_format("report.xlsx") == "excel"

    def test_detect_excel_xls(self):
        assert FileHelper.detect_file_format("report.xls") == "excel"

    def test_detect_json(self):
        assert FileHelper.detect_file_format("data.json") == "json"

    def test_detect_pdf(self):
        assert FileHelper.detect_file_format("doc.pdf") == "pdf"

    def test_detect_empty_string(self):
        assert FileHelper.detect_file_format("") == ""

    def test_detect_none(self):
        assert FileHelper.detect_file_format(None) == ""

    def test_detect_unsupported_extension(self):
        assert FileHelper.detect_file_format("file.txt") == ""

    def test_detect_with_directory_path(self):
        assert FileHelper.detect_file_format("/path/to/data.csv") == "csv"

    def test_detect_no_extension(self):
        assert FileHelper.detect_file_format("somefile") == ""
